"""
@author: Joey Leurs
Version 1.0
Date: 25/07/2022

Webscraping tool to extract price data for the Elder Scrolls Online.
"""

#library
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from playsound import playsound
import pandas as pd
from openpyxl import load_workbook

#webdriver
driver = webdriver.Chrome('E:\Downloads\chromedriver_win32\chromedriver.exe')

#itemlist
df = pd.read_excel('C:/Users/BMO/Desktop/esoprices.xlsx')
writer = pd.ExcelWriter('C:/Users/BMO/Desktop/esoprices.xlsx', engine='openpyxl', mode='a')

pricelist = []

#Scraping
for i, x in enumerate(df['average_link']):
    driver.get(x)
    time.sleep(1)
    averageprice = driver.find_element(By.XPATH, '//*[@id="price-check-result-view"]/div[1]/div/table/tbody/tr[1]/td[3]/span[1]').text
    pricestrip = float(averageprice.replace(',', ''))
    pricelist.append(pricestrip)

#write to excel
pricelist = pd.DataFrame(pricelist)
wb = load_workbook('C:/Users/BMO/Desktop/esoprices.xlsx')
ws = wb['esoprices']
for index, row in pricelist.iterrows():
    cell = 'C%d'  % (index + 2)
    ws[cell] = row[0]
wb.save('C:/Users/BMO/Desktop/esoprices.xlsx')

#eof
print("Prices have been updated")
playsound('C:/Windows/Media/chimes.wav')